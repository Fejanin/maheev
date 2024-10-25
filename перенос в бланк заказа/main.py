import openpyxl
from time import gmtime as gt


current_date = f'{gt().tm_mday:02}.{gt().tm_mon:02}.{gt().tm_year}'


wb = openpyxl.load_workbook('ФА.xlsx', data_only=True)
ws = wb.active

rows = ws.max_row
fed_act = []
for r in range(1, rows + 1):
    fed_act.append(ws.cell(row = r, column = 1).value)
  
file = input('Введите название промежуточного файла: ')

wb = openpyxl.load_workbook(file, data_only=True)
ws = wb.active

rows = ws.max_row
cols = ws.max_column


keys = []
for c in range(1, cols + 1):
    keys.append(ws.cell(row = 1, column = c).value)
    if type(keys[-1]) in [int, float]:
        keys.pop()
        
result = {}
for r in range(2, rows + 1):
    first_key = ''
    if not ws.cell(row = r, column = 1).value:
        continue
    for c, key in enumerate(keys, 1):
        if c == 1:
            first_key = ws.cell(row = r, column = c).value
            result[first_key] = {'FA': False}
        elif key == 'метка':
            if ws.cell(row = r, column = c).value in fed_act:
                result[first_key]['FA'] = True
                result[first_key][key] = ws.cell(row = r, column = c).value
            else:
                result[first_key][key] = ws.cell(row = r, column = c).value
        else:
            result[first_key][key] = ws.cell(row = r, column = c).value

data = {}
for i in result:
    if result[i]['заказ в коробках']:
        data[result[i]['метка']] = (result[i]['FA'], result[i]['заказ в коробках'])


order_file = input('Введите название заводского бланка для заказа: ')

wb = openpyxl.load_workbook(order_file)
ws = wb.active

rows = ws.max_row
cols = ws.max_column

columns = {
    '1 машина': None,
    '2 машина': None,
    }

for r in range(1, rows + 1):
    if all(columns.values()):
        break
    for c in range(1, cols + 1):
        if all(columns.values()):
            break
        cell = ws.cell(row = r, column = c).value
        if cell in columns:
            columns[cell] = c



for r in range(1, rows + 1):
    cell = ws.cell(row = r, column = 1).value
    if cell in data:
        if data[cell][0]:
            ws.cell(row = r, column = columns['1 машина']).value = data[cell][1]
        else:
            ws.cell(row = r, column = columns['2 машина']).value = data[cell][1]
        del data[cell]
wb.save(f'заказ КСК {current_date}.xlsx')

if data:
    ERROR = []
    for i in result:
        if result[i]['метка'] in data:
            ERROR.append(f'{i}:\n\t{result[i]}')

    with open('ERROR.txt', 'w', encoding='utf-8') as f:
        for i in ERROR:
            f.write(f'{i}\n')

