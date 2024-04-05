import openpyxl
from time import gmtime as gt


current_date = f'{gt().tm_mday:02}.{gt().tm_mon:02}.{gt().tm_year}'


def read_table(file):
    name1 = 'бердянск'
    name2 = 'мелитополь'
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active

    rows = ws.max_row
    cols = ws.max_column
    flag = False

    res = {}

    my_dict = {}

    for i in range(1, cols + 1):
        if ws.cell(row = 1, column = i).value:
            if ws.cell(row = 1, column = i).value.lower().strip() == name1:
                col1 = i
            if ws.cell(row = 1, column = i).value.lower().strip() == name2:
                col2 = i


    for i in range(1, rows + 1):
        data = {}
        if not ws.cell(row = i, column = 1).value:
            continue
        if ws.cell(row = i, column = 1).value == 'Номенклатура':
            flag = True
            continue
        if flag:
            for j in [1, col1, col2]:
                cell = ws.cell(row = i, column = j).value
                if cell:
                    if j == col1:
                        data[name1] = cell
                    elif j == col2:
                        data[name2] = cell
                    else:
                        data['СКЮ'] = cell
            if len(data) > 1:
                res[data['СКЮ']] = data
    return res


def read_matrix():
    file = 'матрица.xlsx'
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active

    rows = ws.max_row
    cols = ws.max_column

    d = {}

    for i in range(2, rows + 1):
        key = ws.cell(row = i, column = 1).value
        d[key] = {}
        for j in range(2, cols + 1):
            d[key] = d[key] | {ws.cell(row = 1, column = j).value: ws.cell(row = i, column = j).value}
    return d


def create_res_file(header, orders):
    wb = openpyxl.Workbook()
    ws = wb.active

    all_keys = list(orders.keys())

    for i in range(len(orders) + 1):
        for ind, h in enumerate(header):
            if i == 0:
                ws.cell(row = i + 1, column = ind + 1).value = h
            else:
                if i == 0:
                    ws.cell(row = i + 1, column = ind + 1).value = all_keys[i - 1]
                else:
                    if h in orders[all_keys[i - 1]]:
                        ws.cell(row = i + 1, column = ind + 1).value = orders[all_keys[i - 1]][h]
                    elif h == 'крат. заказа':
                        ws.cell(row = i + 1, column = ind + 1).value = orders[all_keys[i - 1]].get('шт. в уп.', 0) * orders[all_keys[i - 1]].get('кратн. кор.', 0)
                    elif h == 'заказ филиалов':
                        ws.cell(row = i + 1, column = ind + 1).value = orders[all_keys[i - 1]].get('бердянск', 0) + orders[all_keys[i - 1]].get('мелитополь', 0)
                    elif h == 'реком. кратность':
                        krat = ws.cell(row = i + 1, column = header.index('крат. заказа') + 1).value
                        order = ws.cell(row = i + 1, column = header.index('заказ филиалов') + 1).value
                        if order and krat:
                            ws.cell(row = i + 1, column = ind + 1).value = round(order / krat)
                    elif h == 'заказ в коробках':
                        ind_krat_boxes = header.index('кратн. кор.')
                        ind_recom_krat = header.index('реком. кратность')
                        if ind_krat_boxes and ind_recom_krat:
                            ws.cell(row = i + 1, column = ind + 1).value = f'={chr(ord("A") + ind_krat_boxes)}{i + 1}*{chr(ord("A") + ind_recom_krat)}{i + 1}'
                    elif h == 'ОКОНЧАТЕЛЬНЫЙ ЗАКАЗ':
                        ind_order_boxes = header.index('заказ в коробках')
                        ind_count_in_boxes = header.index('шт. в уп.')
                        if ind_order_boxes and ind_count_in_boxes:
                            ws.cell(row = i + 1, column = ind + 1).value = f'={chr(ord("A") + ind_order_boxes)}{i + 1}*{chr(ord("A") + ind_count_in_boxes)}{i + 1}'
                    elif h == 'ВЕС ИТОГО':
                        ind_final_order = header.index('ОКОНЧАТЕЛЬНЫЙ ЗАКАЗ')
                        ind_koof = header.index('кооф')
                        if ind_final_order and ind_koof:
                            ws.cell(row = i + 1, column = ind + 1).value = f'={chr(ord("A") + ind_final_order)}{i + 1}*{chr(ord("A") + ind_koof)}{i + 1}'
    # ws.insert_rows(2)
    # ВЕС ИТОГО
    # =СУММ(M2:M46)
    w_ind = header.index('ВЕС ИТОГО')
    ws.cell(row = 1, column = w_ind + 2).value = f'=SUM({chr(ord("A") + w_ind)}2:{chr(ord("A") + w_ind)}500)'

    wb.save(f'промежуточный расчет {current_date} Махеев.xlsx')



# file = 'тестовый расчет.xlsx'
file = input('Введите название файла: ')
orders = read_table(file)

matrix = read_matrix()

ERROR = {}

for i in orders:
    if i in matrix:
        orders[i] = orders[i] | matrix[i]
    else:
        ERROR[i] = 'Отсутствует в матрице'


header = list(max(orders.values(), key=len).keys()) + ['крат. заказа', 'заказ филиалов', 'реком. кратность', 'заказ в коробках', 'ОКОНЧАТЕЛЬНЫЙ ЗАКАЗ', 'ВЕС ИТОГО']

create_res_file(header, orders)

if ERROR:
    with open('ERROR.txt', 'w', encoding='utf-8') as f:
        for k in ERROR:
            f.write(f'{k} - {ERROR[k]}\n')
